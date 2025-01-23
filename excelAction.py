

import os
import re
import openpyxl
import sys

#修改
def replace_text_in_file(file_path, old_text, new_text):
    with open(file_path, 'r') as file:
        content = file.read()
    content = content.replace(old_text, new_text)
    with open(file_path, 'w') as file:
        file.write(content)

#查找

def find_row(filePath, matchStr):
    with open(filePath, 'r') as file:
        for line in file.readlines():
            if matchStr in line:
                return line.strip('\n')

    

filePath = "/Users/yk/Desktop/jt-station-indonesia-ios/JtStation-Indonesia-iOS/Classes/Manager/Lanaguage/id.lproj/Localizable.strings"
filePath1 = "/Users/yk/Desktop/jt-station-indonesia-ios/JtStation-Indonesia-iOS/Classes/Manager/Lanaguage/en.lproj/Localizable.strings"


# 打开 Excel 文件
workbook = openpyxl.load_workbook('/Users/yk/Downloads/123123123.xlsx')
sheet = workbook.worksheets[0] # 选择工作表（Sheet），默认选择第一个工作表 索引从0开始
def read2output(language, matchs): # language：要提取第几列内容
    # 读取每行的第一项
    count = 0
    for row in sheet.iter_rows(min_row=1, values_only=True):
        if row and row[language]:  # 确保行非空
            print(row[language], row[0])
            oldstr = f"\"{row[2]}\""
            newstr = f"\"{row[language]}\""
            replace_text_in_file(filePath, oldstr, newstr)
            count += 1
    print(count)

def read2output1(): # language：要提取第几列内容
    # 读取每行的第一项
    for row in sheet.iter_rows(min_row=1, values_only=True):
        # print(row)
        if find_row(filePath1, f"\"{row[4]}\"") is not None:
            key = find_row(filePath1, f"\"{row[4]}\"").split('=')[0].strip()
            # print(key)
            oldLine = find_row(filePath,key)
            if oldLine is not None:
                print(oldLine)
                print(row[6])
                newLine = f"{key} = \"{row[6]}\";"
                replace_text_in_file(filePath, oldLine, newLine)

read2output1()
